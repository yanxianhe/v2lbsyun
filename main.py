#!/usr/bin/env python
# -*- encoding: utf-8 -*-
"""
@File    :   main.py
@Time    :   2023/09/16 00:21:58
@Author  :   yxh 
@Version :   1.0
@Contact :   xianhe_yan@sina.com
"""

from fastapi.responses import FileResponse
from loguru import logger
import hashlib
import os
import re
import requests
import urllib3
import logging

from openpyxl import Workbook, load_workbook
from pyproj import Transformer
from pyproj import CRS, Transformer
from coord_convert.transform import wgs2gcj, wgs2bd, gcj2wgs, gcj2bd, bd2wgs, bd2gcj

from fastapi import FastAPI, File, Request, Header, Request, Response, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from modules.db_client import MyredisClient
from public.metadata import Tags

from public.system import Configs
from public.utils import UtilsTools



# # ## ssl
urllib3.disable_warnings()
logging.captureWarnings(True)

description = """"""

# ####################################################################################################
# ##*******************************************************************************************#######
# ##*************************                 os env              *****************************#######
# ##*******************************************************************************************#######
# # ## 提供服务地址
EXTERNAL_URL = os.environ.get("EXTERNAL_URL", "http://10.138.4.163:9095")
# # ## 提供服务地址
HTTPS_TYPE = os.environ.get("HTTPS_TYPE", "https://")
# # ## http 百度源地址
API_WEBGL = os.environ.get("API_WEBGL")

# # ## 百度 ak
API_WEBGL_AK = os.environ.get("API_WEBGL_AK", "API_WEBGL_AK")

# ## 四级域名

# ## 三级域名
MAPOPEN_CDN_BCEBOS_COM = os.environ.get("MAPOPEN_CDN_BCEBOS_COM", "mapopen.cdn.bcebos.com")
API_MAP_BAIDU_COM = os.environ.get("API_MAP_BAIDU_COM", "api.map.baidu.com")
# ## 二级级域名
HM_BAIDU_COM = os.environ.get("HM_BAIDU_COM", "hm.baidu.com")
MAPONLINE0_BDIMG_COM = os.environ.get("MAPONLINE0_BDIMG_COM", "maponline0.bdimg.com")
MAPONLINE1_BDIMG_COM = os.environ.get("MAPONLINE1_BDIMG_COM", "maponline1.bdimg.com")
MAPONLINE2_BDIMG_COM = os.environ.get("MAPONLINE2_BDIMG_COM", "maponline2.bdimg.com")
MAPONLINE3_BDIMG_COM = os.environ.get("MAPONLINE3_BDIMG_COM", "maponline3.bdimg.com")
WEBMAP0_BDIMG_COM = os.environ.get("WEBMAP0_BDIMG_COM", "webmap0.bdimg.com")
PCOR_BAIDU_COM = os.environ.get("PCOR_BAIDU_COM", "pcor.baidu.com")
MIAO_BAIDU_COM = os.environ.get("MIAO_BAIDU_COM", "miao.baidu.com")
DLSWBR_BAIDU_COM = os.environ.get("DLSWBR_BAIDU_COM", "dlswbr.baidu.com")
MAP_BAIDU_COM = os.environ.get("MAP_BAIDU_COM", "map.baidu.com")
# # ## http DMZ 区域代理
if "https" in HTTPS_TYPE:
    # # ## https DMZ 区域代理
    NG_MAPOPEN_CDN_BCEBOS_COM = os.environ.get(
        "NG_MAPOPEN_CDN_BCEBOS_COM", "10.138.4.198:11443/mapopen_cdn_bcebos_com"
    )
    NG_HM_BAIDU_COM = os.environ.get(
        "NG_HM_BAIDU_COM", "10.138.4.198:11443/hm_baidu_com"
    )
    NG_API_MAP_BAIDU_COM = os.environ.get(
        "NG_API_MAP_BAIDU_COM", "10.138.4.198:11443/api_map_baidu_com"
    )
    NG_MAPONLINE0_BDIMG_COM = os.environ.get(
        "NG_MAPONLINE0_BDIMG_COM", "10.138.4.198:11443/maponline0_bdimg_com"
    )
    NG_MAPONLINE1_BDIMG_COM = os.environ.get(
        "NG_MAPONLINE1_BDIMG_COM", "10.138.4.198:11443/maponline1_bdimg_com"
    )
    NG_MAPONLINE2_BDIMG_COM = os.environ.get(
        "NG_MAPONLINE2_BDIMG_COM", "10.138.4.198:11443/maponline2_bdimg_com"
    )
    NG_MAPONLINE3_BDIMG_COM = os.environ.get(
        "NG_MAPONLINE3_BDIMG_COM", "10.138.4.198:11443/maponline3_bdimg_com"
    )
    NG_WEBMAP0_BDIMG_COM = os.environ.get(
        "NG_WEBMAP0_BDIMG_COM", "10.138.4.198:11443/webmap0_bdimg_com"
    )
    NG_PCOR_BAIDU_COM = os.environ.get(
        "NG_PCOR_BAIDU_COM", "10.138.4.198:11443/pcor_baidu_com"
    )
    NG_MIAO_BAIDU_COM = os.environ.get(
        "NG_MIAO_BAIDU_COM", "10.138.4.198:11443/miao_baidu_com"
    )
    NG_DLSWBR_BAIDU_COM = os.environ.get(
        "NG_DLSWBR_BAIDU_COM", "10.138.4.198:11443/dlswbr_baidu_com"
    )
    NG_MAP_BAIDU_COM = os.environ.get(
        "NG_MAP_BAIDU_COM", "10.138.4.198:11443/map_baidu_com"
    )
else:
    NG_MAPOPEN_CDN_BCEBOS_COM = os.environ.get(
        "NG_MAPOPEN_CDN_BCEBOS_COM", "10.138.4.198:11080/mapopen_cdn_bcebos_com"
    )
    NG_HM_BAIDU_COM = os.environ.get(
        "NG_HM_BAIDU_COM", "10.138.4.198:11080/hm_baidu_com"
    )
    NG_API_MAP_BAIDU_COM = os.environ.get(
        "NG_API_MAP_BAIDU_COM", "10.138.4.198:11080/api_map_baidu_com"
    )
    NG_MAPONLINE0_BDIMG_COM = os.environ.get(
        "NG_MAPONLINE0_BDIMG_COM", "10.138.4.198:11080/maponline0_bdimg_com"
    )
    NG_MAPONLINE1_BDIMG_COM = os.environ.get(
        "NG_MAPONLINE1_BDIMG_COM", "10.138.4.198:11080/maponline1_bdimg_com"
    )
    NG_MAPONLINE2_BDIMG_COM = os.environ.get(
        "NG_MAPONLINE2_BDIMG_COM", "10.138.4.198:11080/maponline2_bdimg_com"
    )
    NG_MAPONLINE3_BDIMG_COM = os.environ.get(
        "NG_MAPONLINE3_BDIMG_COM", "10.138.4.198:11080/maponline3_bdimg_com"
    )
    NG_WEBMAP0_BDIMG_COM = os.environ.get(
        "NG_WEBMAP0_BDIMG_COM", "10.138.4.198:11080/webmap0_bdimg_com"
    )
    NG_PCOR_BAIDU_COM = os.environ.get(
        "NG_PCOR_BAIDU_COM", "10.138.4.198:11080/pcor_baidu_com"
    )
    NG_MIAO_BAIDU_COM = os.environ.get(
        "NG_MIAO_BAIDU_COM", "10.138.4.198:11080/miao_baidu_com"
    )
    NG_DLSWBR_BAIDU_COM = os.environ.get(
        "NG_DLSWBR_BAIDU_COM", "10.138.4.198:11080/dlswbr_baidu_com"
    )
    NG_MAP_BAIDU_COM = os.environ.get(
        "NG_MAP_BAIDU_COM", "10.138.4.198:11080/map_baidu_com"
    )

if API_WEBGL is not None:
    _webgl_ = API_WEBGL
else:
    _webgl_ = Configs._WEBGL_


#####################################################################################################
###*******************************************************************************************#######
###*************************             FastAPI init            *****************************#######
###*******************************************************************************************#######
try:
    tags_metadata = Tags.tags_metadata()
    app = FastAPI(
        title="jsapiAPI",
        description=description,
        summary="",
        openapi_url="/jsapiAPI.json",
        version="v1.0.0",
        openapi_tags=tags_metadata,
    )

    app.add_middleware(
        CORSMiddleware,
        allow_origins=["*"],
        allow_credentials=True,
        allow_methods=["*"],
        allow_headers=["*"],
    )
    logger.success("[success] running (Press CTRL+C to quit) ")
except Exception as e:
    logger.error("[error] running  %s" % e)
    raise e


    
# ####################################################################################################
# ##*******************************************************************************************#######
# ##*************************              JSAPI init             *****************************#######
# ##*******************************************************************************************#######

async def send_http(_uuid, _requrl, _headers, payload={}):
    try:
        response = requests.request(
            "GET", _requrl, headers=_headers, data=payload, verify=False, timeout=10
        )
        # 设置响应编码为 UTF-8
        # response.encoding = 'UTF-8'
        # 获取响应的估计编码方式
        encoding = response.apparent_encoding
        # 设置响应编码为估计编码方式
        response.encoding = encoding
    except requests.exceptions.Timeout:
        msg = "Internal#Server#Error TimeoutException [%s] " % _requrl
        logging.error(("[%s] {%s}" % (_uuid, msg)))
        return msg
    except requests.exceptions.ConnectionError as e:
        msg = "Internal#Server#Error Exception [%s] " % _requrl
        logging.error(("[%s] {%s}" % (_uuid, msg)))
        return "Internal#Server#Error Exception %s" % e
    except requests.exceptions.HTTPError as e:
        logging.error(("[%s] {%s}" % (_uuid, msg)))
        return "Internal#Server#Error Exception 0 %s" % e
    return response.text
# 清理key
def clear_redis_key(ak):
    MyredisClient.delete_redis_data(("%s%s") % (Configs._JSAPI_KEY, ak))
    MyredisClient.delete_redis_data(("%s%s") % (Configs._GETSCRIPT_KEY, ak))
    MyredisClient.delete_redis_data(("%s%s") % (Configs._GETBMAPCSS_KEY, ak))
    MyredisClient.delete_redis_data(("%s%s") % (Configs._JSAPI_NEW_KEY, ak))
    MyredisClient.delete_redis_data(("%s%s") % (Configs._GETSCRIPT_NEW_KEY, ak))
    MyredisClient.delete_redis_data(("%s%s") % (Configs._GETBMAPCSS_NEW_KEY, ak))

async def function_url(ak, srt):
    # # ## 源 js url
    _webgl_js = ""
    # # ## 源 css url
    _webgl_css = ""
    _key_new = ("%s%s") % (Configs._JSAPI_NEW_KEY, ak)
    _functon_webgl_new = ""
    # 正则表达式匹配模式 查找符合模式的URL
    pattern = re.compile(r'https?://[^\s"\']+')
    urls = re.findall(pattern, srt)
    for url in urls:
        if "getscript" in url.lower():
            if ".baidu.com" in _webgl_.lower() :
                _webgl_js = url
            else :
                _rsync_webgl_js = getscriot_new(url)
                _webgl_js_tmp = await _rsync_webgl_js
                _webgl_js = _webgl_js_tmp.replace("https://", "http://")
            _functon_webgl_new = srt.replace(
                url, "%s/v2lbsyun/getscript/%s" % (EXTERNAL_URL, ak)
            )
        if ".css" in url.lower():
                if ".baidu.com" in _webgl_.lower() :
                    _webgl_css = url
                else :
                    _rsync_webgl_css = getscriot_new(url)
                    _webgl_css_tmp = await _rsync_webgl_css
                    _webgl_css = _webgl_css_tmp.replace("https://", "http://")
                    
                _functon_webgl_new = _functon_webgl_new.replace(
                url, "%s/v2lbsyun/getbmapcss/%s" % (EXTERNAL_URL, ak)
            )
    # ## 拼接后的 function 放到redis
    MyredisClient.pull_redis_data(_key_new, _functon_webgl_new)

    # ## js key
    _getscript_key = ("%s%s") % (Configs._GETSCRIPT_KEY, ak)
    _getscript_new_key = ("%s%s") % (Configs._GETSCRIPT_NEW_KEY, ak)
    # ## css key
    _getbmapcss_key = ("%s%s") % (Configs._GETBMAPCSS_KEY, ak)
    _getbmapcss_new_key = ("%s%s") % (Configs._GETBMAPCSS_NEW_KEY, ak)

    _uuid = UtilsTools().getUuid1()
    # ## 获取源js
    _rsync_getscript = send_http(_uuid, _webgl_js, _headers={})
    _getscript = await _rsync_getscript
    if "Internal#Server#Error" in _getscript:
        logger.error(("[%s] {%s}" % ("ER99999:请求互联网失败!", _webgl_js)))
    else:
        logger.info(_getscript)
        # ## 获取源js 入库
        MyredisClient.pull_redis_data(_getscript_key, _getscript)
        # ## 替换后的 js
        _rsync_getscript_tmp = getscriot_new(_getscript)
        _getscript_tmp = await _rsync_getscript_tmp
        MyredisClient.pull_redis_data(_getscript_new_key, _getscript_tmp)
    # ## 获取源css 入库
    _rsync_getbmapcss = send_http(_uuid, _webgl_css, _headers={})
    _getbmapcss = await _rsync_getbmapcss
    if "Internal#Server#Error" in _getscript:
        logger.error(("[%s] {%s}" % ("ER99999:请求互联网失败!", _webgl_css)))
    else:
        logger.debug(_getbmapcss)
        # ## 获取源css 入库
        MyredisClient.pull_redis_data(_getbmapcss_key, _getbmapcss)
        # ## 替换后的 css

        _rsync_getbmapcss_tmp = getscriot_new(_getbmapcss)
        _getbmapcss_tmp = await _rsync_getbmapcss_tmp
        MyredisClient.pull_redis_data(_getbmapcss_new_key, _getbmapcss_tmp)
    return _functon_webgl_new


# 替换js css
async def getscriot_new(srt):
    # ## 替换域名
    tmp_srt = (
        srt.replace(MAPOPEN_CDN_BCEBOS_COM, NG_MAPOPEN_CDN_BCEBOS_COM, -1)
        .replace(HM_BAIDU_COM, NG_HM_BAIDU_COM, -1)
        .replace(API_MAP_BAIDU_COM, NG_API_MAP_BAIDU_COM, -1)
        .replace(MAPONLINE0_BDIMG_COM, NG_MAPONLINE0_BDIMG_COM, -1)
        .replace(MAPONLINE1_BDIMG_COM, NG_MAPONLINE1_BDIMG_COM, -1)
        .replace(MAPONLINE2_BDIMG_COM, NG_MAPONLINE2_BDIMG_COM, -1)
        .replace(MAPONLINE3_BDIMG_COM, NG_MAPONLINE3_BDIMG_COM, -1)
        .replace(WEBMAP0_BDIMG_COM, NG_WEBMAP0_BDIMG_COM, -1)
        .replace(PCOR_BAIDU_COM, NG_PCOR_BAIDU_COM, -1)
        .replace(MIAO_BAIDU_COM, NG_MIAO_BAIDU_COM, -1)
        .replace(DLSWBR_BAIDU_COM, NG_DLSWBR_BAIDU_COM, -1)
        .replace(MAP_BAIDU_COM, NG_MAP_BAIDU_COM, -1)
    )
    logger.info(tmp_srt)
    return tmp_srt
# 要将大地2000坐标转换为百度坐标（BD-09）
def cgcs2000bd09(latitude, longitude):
    if latitude is not None and latitude != "" and longitude is not None and longitude != "":
        # 定义大地2000坐标
        # 创建转换器
        # 创建转换器
        transformer = Transformer.from_crs('EPSG:4326', 'EPSG:3857')  # 大地2000坐标 -> Web墨卡托投影坐标
        bd_transformer = Transformer.from_crs('EPSG:3857','EPSG:4490') # Web墨卡托投影坐标 -> BD-09坐标

        # 坐标转换
        web_mercator_x, web_mercator_y = transformer.transform(latitude, longitude)  # 大地2000坐标 -> Web墨卡托投影坐标
        bd_longitude, bd_latitude = bd_transformer.transform(web_mercator_y, web_mercator_x)  # Web墨卡托投影坐标 -> BD-09坐标

        # 打印结果
        logger.debug(f"BD-09坐标: {bd_longitude}, {bd_latitude}")
        return bd_latitude , bd_longitude
    else:
        return latitude, longitude

# lat lon
def crs_cgcs2000(east, north) :
    if east is not None and east != "" and north is not None and north != "":
        # 定义CGCS2000投影坐标系
        crs_cgcs2000 = CRS.from_epsg(4545)
        # 创建坐标转换器
        transformer_cgcs_wgs = Transformer.from_crs(crs_cgcs2000, 'epsg:4326')  # CGCS2000 to WGS84
        transformer_wgs_cgcs = Transformer.from_crs('epsg:4326', crs_cgcs2000)  # WGS84 to CGCS2000

        # CGCS2000坐标转换为WGS84经纬度
        # east, north =4349515.286,623077.749 # 假设CGCS2000坐标为 (1000000, 2000000)
        lon, lat = transformer_cgcs_wgs.transform(east, north)
        # WGS84 to BD09
        lon_bd09, lat_bd09 = wgs2bd(lon, lat )
        return lon_bd09, lat_bd09
    else :
        return east, north
# 处理函数示例
@app.get("/v1/ping", tags=["ping"])
async def ping_controllers(request: Request):
    # 查询redis数据
    return {"message": request.url, "nu": MyredisClient.dbsize_redis()}


# ####################################################################################################

# jsapi getscript getbmapcss os.environ.get('xx', '127.0.0.1')

# https://api.map.baidu.com/api?v=1.0&type=webgl&ak=


@app.get("/v2lbsyun/webgl/{ak}", tags=["jsapi"])
async def jsapi_controllers(ak, response: Response, request: Request):
    _uuid = UtilsTools().getUuid1()
    logger.debug(("[%s] {%s}" % (_uuid, ak)))
    # # ## 查询redis key 是否存在如果存在直接从 redis 中获取
    _key = ("%s%s") % (Configs._JSAPI_KEY, ak)
    _key_new = ("%s%s") % (Configs._JSAPI_NEW_KEY, ak)
    _redis_data = MyredisClient.get_redis_data(_key_new)

    if _redis_data is not None and _redis_data != "" and len(_redis_data) > 0:
        return Response(content=_redis_data, media_type="text/plain")
    else:
        # ## 清理key ##
        clear_redis_key(ak)
        logger.info(("[%s] {%s}" % (_uuid, "clear_redis_key")))
        # ## 通过接口获取方法
        _requrl = ("%s%s") % (_webgl_, ak)
        logger.debug(("[%s] {%s}" % (_uuid, _requrl)))
        logger.info("[%s][%s][%s]  " % (_uuid, request.client.host, request.url))
        # ## 源方法
        _rsync_functon_webgl = send_http(_uuid, _requrl, _headers={})
        _functon_webgl = await _rsync_functon_webgl
  
        # ## 接口是否异常异常直接返回给前端
        if "Internal#Server#Error" in _functon_webgl:
            return _functon_webgl
        else:
            MyredisClient.pull_redis_data(_key, _functon_webgl)
            if (
                _functon_webgl is not None
                and _functon_webgl != ""
                and len(_functon_webgl) > 0
            ):
                _rsync_function_url=function_url(ak, _functon_webgl)
                _functon_webgl_new = await _rsync_function_url

                return Response(content=_functon_webgl_new, media_type="text/plain")
            else:
                return {"message": "ER88888:处理数据异常非标准数据!", "url": _requrl}


@app.get("/v2lbsyun/getscript/{ak}", tags=["getscript"])
async def getscript_controllers(ak, response: Response, request: Request):
    # 直接返回
    _key = ("%s%s") % (Configs._GETSCRIPT_NEW_KEY, ak)
    _redis_data = MyredisClient.get_redis_data(_key)
    if _redis_data is not None:
        return Response(content=_redis_data, media_type="text/plain")
    else:
        _key = ("%s%s") % (Configs._JSAPI_NEW_KEY, ak)
        MyredisClient.delete_redis_data(_key)
        return {"message": ("%s%s") % (_webgl_, ak)}


@app.get("/v2lbsyun/getbmapcss/{ak}", tags=["getbmapcss"])
async def getbmapcss_controllers(ak, response: Response, request: Request):
    # 如果有直接返回
    _key = ("%s%s") % (Configs._GETBMAPCSS_NEW_KEY, ak)
    _redis_data = MyredisClient.get_redis_data(_key)
    if _redis_data is not None:
        return Response(content=_redis_data, media_type="text/plain")
    else:
        _key = ("%s%s") % (Configs._JSAPI_NEW_KEY, ak)
        MyredisClient.delete_redis_data(_key)
        return {"message": ("%s%s") % (_webgl_, ak)}


# 处理函数示例
@app.get("/v2lbsyun/clear/{ak}", tags=["clear"])
async def clear_controllers(ak, sk, response: Response, request: Request):
    _uuid = UtilsTools().getUuid1()
    logger.info("[%s][%s][%s]  " % (_uuid, request.client.host, request.url))
    #  计算 32 小写
    hash_object2 = hashlib.md5(
        hashlib.md5(ak.encode()).hexdigest().encode()
    ).hexdigest()
    if sk in hash_object2:
        clear_redis_key(ak)
        logger.info("[%s][%s][%s]" % (request.url, _uuid, request.client.host))
        return {"message": request.url}
    else:
        logger.info("[%s][%s][%s]" % (request.url, _uuid, request.client.host))
        return {"message": "认证失败！"}


# 处理函数示例

@app.post("/v2lbsyun/crscgcs2000eastnorth/{ak}", tags=["crscgcs2000"])
async def crscgcs2000eastnorth(response: Response, request: Request,file: UploadFile = File(...)):

    _tmp = "%s-%s"%(hashlib.md5(file.filename.encode()).hexdigest().encode(),file.filename)
    _filename = _tmp[2:]
    # 读取上传的 xlsx 文件数据
    content = await file.read()
    with open(_filename, "wb") as f:
        f.write(content)

    # 打开上传的 xlsx 文件
    wb = load_workbook(_filename)
    # 读取工作表数据（示例中只读取第一个工作表）
    sheet = wb.active
    # 修改第三 四列的内容
    for row in sheet.iter_rows(min_row=2, min_col=3, max_col=4):
        latitude, longitude = crs_cgcs2000(row[0].value, row[1].value)
        # latitude, longitude = crs_cgcs2000(row[1].value, row[0].value)
        row[0].value = latitude
        row[1].value = longitude

    # 保存修改后的文件
    modified_filename = UtilsTools().getUuid1()+".xlsx"
    wb.save(modified_filename)

    # 返回修改后的文件下载响应
    return FileResponse(modified_filename, filename=modified_filename)




@app.post("/v2lbsyun/crscgcs2000northeast/{ak}", tags=["crscgcs2000"])
async def crscgcs2000northeast(response: Response, request: Request,file: UploadFile = File(...)):

    _tmp = "%s-%s"%(hashlib.md5(file.filename.encode()).hexdigest().encode(),file.filename)
    _filename = _tmp[2:]
    # 读取上传的 xlsx 文件数据
    content = await file.read()
    with open(_filename, "wb") as f:
        f.write(content)

    # 打开上传的 xlsx 文件
    wb = load_workbook(_filename)
    # 读取工作表数据（示例中只读取第一个工作表）
    sheet = wb.active
    # 修改第三 四列的内容
    for row in sheet.iter_rows(min_row=2, min_col=3, max_col=4):
        #latitude, longitude = crs_cgcs2000(row[0].value, row[1].value)
        latitude, longitude = crs_cgcs2000(row[1].value, row[0].value)
        row[0].value = latitude
        row[1].value = longitude

    # 保存修改后的文件
    modified_filename = UtilsTools().getUuid1()+".xlsx"
    wb.save(modified_filename)

    # 返回修改后的文件下载响应
    return FileResponse(modified_filename, filename=modified_filename)